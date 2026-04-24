import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
import sqlite3
from config import DB_FILE, OUTPUT_FILE
def export_to_excel():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    
    workbook = openpyxl.Workbook()
    _write_books_sheet(workbook, conn)
    _write_summary_sheet(workbook, conn)

    conn.close()

    workbook.save(OUTPUT_FILE)
    print(f"\nExcel file saved: {OUTPUT_FILE}")
def _write_books_sheet(workbook: openpyxl.Workbook, conn: sqlite3.Connection):
    """Write all book records to the first sheet."""

    sheet = workbook.active
    sheet.title = "Books"

    headers = ["#", "Title", "Author", "Publish Year", "Subject", "ISBN"]
    _write_header_row(sheet, headers)

    rows = conn.execute("""
        SELECT title, author, publish_year, subject, isbn
        FROM books
        ORDER BY subject, title
    """).fetchall()

    for index, row in enumerate(rows, start=1):
        sheet.append([
            index,
            row["title"],
            row["author"],
            row["publish_year"],
            row["subject"],
            row["isbn"],
        ])

    _style_data_rows(sheet, start_row=2, total_rows=len(rows))
    _set_column_widths(sheet, [5, 50, 30, 15, 15, 20])
    sheet.freeze_panes = "A2"


def _write_summary_sheet(workbook: openpyxl.Workbook, conn: sqlite3.Connection):
    """Write a per-subject record count to the second sheet."""

    sheet = workbook.create_sheet(title="Summary")

    headers = ["Subject", "Record Count"]
    _write_header_row(sheet, headers)

    rows = conn.execute("""
        SELECT subject, COUNT(*) as count
        FROM books
        GROUP BY subject
        ORDER BY count DESC
    """).fetchall()

    for row in rows:
        sheet.append([row["subject"], row["count"]])

    _style_data_rows(sheet, start_row=2, total_rows=len(rows))
    _set_column_widths(sheet, [20, 15])
    sheet.freeze_panes = "A2"


def _write_header_row(sheet, headers: list[str]):
    """Write and style the header row."""

    sheet.append(headers)
    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )
    header_font = Font(
        name="Calibri",
        bold=True,
        color="FFFFFF",
        size=11
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.row_dimensions[1].height = 20


def _style_data_rows(sheet, start_row: int, total_rows: int):
    """Apply alternating row colors to data rows."""

    light_blue = PatternFill(
        start_color="DCE6F1",
        end_color="DCE6F1",
        fill_type="solid"
    )
    white = PatternFill(fill_type=None)

    for row_num in range(start_row, start_row + total_rows):
        fill = light_blue if row_num % 2 == 0 else white
        for cell in sheet[row_num]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")


def _set_column_widths(sheet, widths: list[int]):
    """Set explicit column widths."""
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width